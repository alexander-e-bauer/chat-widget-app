import pandas as pd
import numpy as np
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime, timedelta
import calendar


class MerrillScorecardManager:
    def __init__(self):
        self.wb = Workbook()
        self.metrics = {
            'Growth_Award': [
                'Net New Households',
                'Net New Money Flow',
                'Revenue Growth',
                'Investment Solutions Adoption',
                'Digital Activation Rate'
            ],
            'Team_Growth_Award': [
                'Team Revenue Growth',
                'Team Client Acquisition',
                'Team Solutions Adoption',
                'Team Collaboration Score',
                'Digital Team Engagement'
            ]
        }

    def create_scorecard(self, output_file):
        # Remove default sheet
        self.wb.remove(self.wb.active)

        # Create main tracking sheets
        self._create_lpb_sheet()
        self._create_nr_sheet()
        self._create_pipeline_sheet()
        self._create_pmac_sheet()

        # Apply formatting and save
        self.wb.save(output_file)
        print(f"Scorecard created successfully: {output_file}")

    def _create_lpb_sheet(self):
        ws = self.wb.create_sheet("LPB Private Wealth")
        self._setup_tracking_sheet(ws, "LPB Private Wealth Advisors")

    def _create_nr_sheet(self):
        ws = self.wb.create_sheet("NR Wealth Management")
        self._setup_tracking_sheet(ws, "NR Wealth Management Advisors")

    def _setup_tracking_sheet(self, ws, title):
        # Header styling
        header_fill = PatternFill(start_color="041E42", end_color="041E42", fill_type="solid")  # Merrill Blue
        header_font = Font(color="FFFFFF", bold=True)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # Title
        ws.merge_cells('A1:K1')
        ws['A1'] = title
        ws['A1'].font = Font(size=14, bold=True)
        ws['A1'].alignment = Alignment(horizontal='center')

        # Column headers
        headers = ['Advisor Name', 'Current Week', 'YTD Progress']
        for metric in self.metrics['Growth_Award']:
            headers.extend([f'{metric} Target', f'{metric} Actual'])

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=2, column=col)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.border = border

        # Set column widths
        ws.column_dimensions['A'].width = 20
        for col in range(2, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col)].width = 15

    def _create_pipeline_sheet(self):
        ws = self.wb.create_sheet("90-Day Pipeline")

        # Pipeline sheet headers
        headers = [
            'Advisor Name',
            'Client Name',
            'Opportunity Type',
            'Potential Value',
            'Probability',
            'Expected Close Date',
            'Impact Metric',
            'Current Status',
            'Next Steps',
            'Last Updated'
        ]

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col)
            cell.value = header
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="041E42", end_color="041E42", fill_type="solid")
            cell.font = Font(color="FFFFFF", bold=True)
            ws.column_dimensions[get_column_letter(col)].width = 15

    def _create_pmac_sheet(self):
        ws = self.wb.create_sheet("PMAC Results")

        # Create sections for both award types
        current_row = 1
        for award_type, metrics in self.metrics.items():
            # Award type header
            ws.merge_cells(f'A{current_row}:E{current_row}')
            ws[f'A{current_row}'] = award_type
            ws[f'A{current_row}'].font = Font(size=12, bold=True)
            ws[f'A{current_row}'].fill = PatternFill(start_color="041E42", end_color="041E42", fill_type="solid")
            ws[f'A{current_row}'].font = Font(color="FFFFFF", bold=True)

            # Headers
            current_row += 1
            headers = ['Metric', 'Target', 'Current', 'Progress', 'Status']
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=current_row, column=col)
                cell.value = header
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="E1E1E1", end_color="E1E1E1", fill_type="solid")

            # Metrics
            for metric in metrics:
                current_row += 1
                ws[f'A{current_row}'] = metric

            current_row += 2  # Space between award types

        # Set column widths
        ws.column_dimensions['A'].width = 30
        for col in ['B', 'C', 'D', 'E']:
            ws.column_dimensions[col].width = 15

    def update_advisor_metrics(self, sheet_name, advisor_data):
        """
        Update advisor metrics in the specified sheet

        advisor_data = {
            'name': 'Advisor Name',
            'metrics': {
                'Net New Households': {'target': 10, 'actual': 8},
                ...
            }
        }
        """
        ws = self.wb[sheet_name]
        # Implementation for updating advisor metrics
        pass

    def update_pipeline(self, pipeline_data):
        """
        Update the 90-day pipeline sheet

        pipeline_data = [{
            'advisor_name': 'Name',
            'client_name': 'Client',
            'opportunity_type': 'Type',
            ...
        }]
        """
        ws = self.wb["90-Day Pipeline"]
        # Implementation for updating pipeline
        pass

    def update_pmac_results(self, pmac_data):
        """
        Update PMAC results

        pmac_data = {
            'Growth_Award': {
                'Net New Households': {'target': 100, 'current': 75},
                ...
            },
            'Team_Growth_Award': {
                ...
            }
        }
        """
        ws = self.wb["PMAC Results"]
        # Implementation for updating PMAC results
        pass


def main():
    # Create scorecard manager instance
    manager = MerrillScorecardManager()

    # Generate initial scorecard
    output_file = f'Merrill_Scorecard_{datetime.now().strftime("%Y%m%d")}.xlsx'
    manager.create_scorecard(output_file)

    # Example of how to update data (commented out as example only)
    """
    # Update advisor metrics
    advisor_data = {
        'name': 'John Smith',
        'metrics': {
            'Net New Households': {'target': 10, 'actual': 8},
            'Net New Money Flow': {'target': 1000000, 'actual': 750000},
        }
    }
    manager.update_advisor_metrics('LPB Private Wealth', advisor_data)

    # Update pipeline
    pipeline_data = [{
        'advisor_name': 'John Smith',
        'client_name': 'Client A',
        'opportunity_type': 'New Money',
        'potential_value': 500000,
        'probability': 0.75,
        'expected_close_date': '2024-12-31',
        'impact_metric': 'Net New Money Flow',
        'current_status': 'In Progress',
        'next_steps': 'Follow-up meeting scheduled',
        'last_updated': '2024-11-25'
    }]
    manager.update_pipeline(pipeline_data)
    """


if __name__ == "__main__":
    main()