import os
import logging
from datetime import datetime
import json
from typing import Dict, List, Any, Optional
import openai
from dotenv import load_dotenv
import pdfplumber
import io
from openpyxl import Workbook, load_workbook
from openpyxl.styles import (
    PatternFill, Border, Side, Alignment, Protection, Font,
    NamedStyle, Color, numbers
)
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.dimensions import ColumnDimension
from openpyxl.worksheet.table import Table, TableStyleInfo

# Load environment variables
load_dotenv()

# Initialize OpenAI client
client = openai.OpenAI(api_key=os.getenv('OPENAI_API_KEY'))


class ExcelStyler:
    """Helper class for Excel styling"""

    @staticmethod
    def create_header_style():
        header_style = NamedStyle(name="header_style")
        header_style.font = Font(bold=True, color="FFFFFF", size=12)
        header_style.fill = PatternFill("solid", fgColor="2F75B5")
        header_style.alignment = Alignment(horizontal="center", vertical="center")
        header_style.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        return header_style

    @staticmethod
    def create_currency_style():
        currency_style = NamedStyle(name="currency_style")
        currency_style.number_format = '"$"#,##0.00'
        currency_style.font = Font(size=11)
        currency_style.alignment = Alignment(horizontal="right")
        return currency_style

    @staticmethod
    def create_date_style():
        date_style = NamedStyle(name="date_style")
        date_style.number_format = 'YYYY-MM-DD'
        date_style.font = Font(size=11)
        date_style.alignment = Alignment(horizontal="center")
        return date_style

    @staticmethod
    def create_number_style():
        number_style = NamedStyle(name="number_style")
        number_style.number_format = '#,##0.00'
        number_style.font = Font(size=11)
        number_style.alignment = Alignment(horizontal="right")
        return number_style


import logging
import os


class PaystubProcessor:
    def __init__(self, output_file: str = "financial_master.xlsx"):
        self.output_file = output_file
        self.logger = self._setup_logger()  # This line needs the method below

        # Define headers as class attribute
        self.headers = [
            'Date', 'Pay Period Start', 'Pay Period End', 'Gross Pay',
            'Net Pay', 'Regular Hours', 'Overtime Hours', 'Regular Earnings',
            'Overtime Earnings', 'Total Taxes', 'Federal Tax', 'State Tax',
            'Social Security', 'Medicare', 'Pre-tax Deductions',
            'Post-tax Deductions', 'Total Deductions'
        ]

        self.initialize_excel()

    def _setup_logger(self):
        """Set up and return a logger for the class"""
        # Create logger
        logger = logging.getLogger('budget')
        logger.setLevel(logging.INFO)

        # Create handlers
        console_handler = logging.StreamHandler()
        console_handler.setLevel(logging.INFO)

        # Create log directory if it doesn't exist
        log_dir = 'logs'
        if not os.path.exists(log_dir):
            os.makedirs(log_dir)

        # Create file handler
        file_handler = logging.FileHandler(os.path.join(log_dir, 'budget.log'))
        file_handler.setLevel(logging.INFO)

        # Create formatters
        console_formatter = logging.Formatter('%(levelname)s - %(message)s')
        file_formatter = logging.Formatter('%(asctime)s - %(name)s - %(levelname)s - %(message)s')

        # Add formatters to handlers
        console_handler.setFormatter(console_formatter)
        file_handler.setFormatter(file_formatter)

        # Add handlers to logger
        logger.addHandler(console_handler)
        logger.addHandler(file_handler)

        return logger

    def initialize_excel(self):
        """Initialize the Excel file with formatting if it doesn't exist"""
        if not os.path.exists(self.output_file):
            wb = Workbook()
            ws = wb.active
            ws.title = "Paystub Data"

            # Create and register styles
            styles = {
                'header': ExcelStyler.create_header_style(),
                'currency': ExcelStyler.create_currency_style(),
                'date': ExcelStyler.create_date_style(),
                'number': ExcelStyler.create_number_style()
            }

            for style in styles.values():
                if style.name not in wb.named_styles:
                    wb.add_named_style(style)

            # Write headers and set column widths
            for col, header in enumerate(self.headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.style = "header_style"
                ws.column_dimensions[get_column_letter(col)].width = 15

            self._create_or_update_table(ws, 1)  # Initialize with just headers
            wb.save(self.output_file)
            self.logger.info(f"Initialized new Excel file: {self.output_file}")

    def _create_or_update_table(self, worksheet, last_row: int):
        """Create or update the Excel table"""
        table_name = "PaystubData"

        # Remove existing table if it exists
        if table_name in worksheet.tables:
            del worksheet.tables[table_name]

        # Create new table
        tab = Table(
            displayName=table_name,
            ref=f"A1:{get_column_letter(len(self.headers))}{last_row}"
        )

        # Set table style
        tab.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium9",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False
        )

        worksheet.add_table(tab)

    def process_pdf(self, pdf_data: bytes) -> Optional[Dict[str, Any]]:
        """Process PDF data and extract paystub information using GPT"""
        try:
            # Convert bytes to text using pdfplumber
            pdf_file = io.BytesIO(pdf_data)
            with pdfplumber.open(pdf_file) as pdf:
                text = "\n".join(page.extract_text() for page in pdf.pages)

            # Process text with GPT
            paystub_data = self.process_paystub_with_gpt(text)
            if not paystub_data:
                self.logger.error("Failed to extract data from PDF")
                return None

            # Save to Excel
            self.save_to_excel(paystub_data)
            return paystub_data

        except Exception as e:
            self.logger.error(f"Error processing PDF: {str(e)}")
            return None

    def process_paystub_with_gpt(self, text: str) -> Dict[str, Any]:
        """Process paystub text using OpenAI's GPT model"""
        prompt = """
        Extract the following information from the paystub text in JSON format:
        - Pay period start date
        - Pay period end date
        - Pay date
        - Gross pay
        - Net pay
        - Regular hours worked
        - Regular earnings
        - Total taxes
        - Federal tax withholding
        - State tax withholding
        - Social security tax
        - Medicare tax
        - Pre-tax deductions total
        - Post-tax deductions total

        Return the data in this format:
        {
            "pay_period_start": "YYYY-MM-DD",
            "pay_period_end": "YYYY-MM-DD",
            "pay_date": "YYYY-MM-DD",
            "gross_pay": 0000.00,
            "net_pay": 0000.00,
            "regular_hours": 00.00,
            "regular_earnings": 0000.00,
            "total_taxes": 0000.00,
            "federal_tax": 0000.00,
            "state_tax": 0000.00,
            "social_security": 0000.00,
            "medicare": 0000.00,
            "pre_tax_deductions": 0000.00,
            "post_tax_deductions": 0000.00
        }
        """

        try:
            response = client.chat.completions.create(
                model="gpt-4",
                messages=[
                    {"role": "system",
                     "content": "You are a financial data extraction assistant. Extract the requested information from the paystub and return it in JSON format."},
                    {"role": "user", "content": f"{prompt}\n\nPaystub text:\n{text}"}
                ]
            )

            return json.loads(response.choices[0].message.content)
        except Exception as e:
            self.logger.error(f"Error processing with GPT: {str(e)}")
            raise

    def save_to_excel(self, paystub_data: Dict[str, Any]):
        """Save processed paystub data to the Excel file"""
        try:
            wb = load_workbook(self.output_file)
            ws = wb.active

            # Prepare row data
            row_data = [
                paystub_data['pay_date'],
                paystub_data['pay_period_start'],
                paystub_data['pay_period_end'],
                paystub_data['gross_pay'],
                paystub_data['net_pay'],
                paystub_data['regular_hours'],
                0,  # overtime_hours
                paystub_data['regular_earnings'],
                0,  # overtime_earnings
                paystub_data['total_taxes'],
                paystub_data['federal_tax'],
                paystub_data['state_tax'],
                paystub_data['social_security'],
                paystub_data['medicare'],
                paystub_data['pre_tax_deductions'],
                paystub_data['post_tax_deductions'],
                paystub_data['pre_tax_deductions'] + paystub_data['post_tax_deductions']
            ]

            # Find the last data row (excluding summary)
            last_row = ws.max_row
            # Check if summary exists (look for "Summary" in first column)
            while last_row > 0 and ws.cell(row=last_row, column=1).value == "Summary":
                last_row -= 1

            # Insert new data row
            next_row = last_row + 1

            # Write data with appropriate styling
            for col, value in enumerate(row_data, 1):
                cell = ws.cell(row=next_row, column=col, value=value)

                if col in [1, 2, 3]:  # Date columns
                    cell.style = "date_style"
                elif col in [4, 5, 8, 9, 10, 11, 12, 13, 14, 15, 16, 17]:  # Money columns
                    cell.style = "currency_style"
                else:  # Number columns
                    cell.style = "number_style"

            # Update the table to include the new row
            self._create_or_update_table(ws, next_row)

            # Add or update summary section (always 2 rows below the table)
            summary_row = next_row + 2

            # Add summary
            ws.cell(row=summary_row, column=1, value="Summary").style = "header_style"
            ws.cell(row=summary_row, column=2, value="Total Gross Pay").style = "header_style"
            ws.cell(row=summary_row, column=3, value=f"=SUM(D2:D{next_row})").style = "currency_style"
            ws.cell(row=summary_row, column=4, value="Total Net Pay").style = "header_style"
            ws.cell(row=summary_row, column=5, value=f"=SUM(E2:E{next_row})").style = "currency_style"

            wb.save(self.output_file)
            self.logger.info(f"Successfully saved data to {self.output_file}")

        except Exception as e:
            self.logger.error(f"Error saving to Excel: {str(e)}")
            raise


def main():
    processor = PaystubProcessor()

    try:
        # Example usage with PDF data
        with open("paystubs/Paystub2.pdf", "rb") as pdf_file:
            pdf_data = pdf_file.read()
            result = processor.process_pdf(pdf_data)

            if result:
                print("Paystub processed successfully!")
            else:
                print("Failed to process paystub.")

    except Exception as e:
        print(f"Error: {str(e)}")


if __name__ == "__main__":
    main()